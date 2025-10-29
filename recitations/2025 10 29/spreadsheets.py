from openpyxl import Workbook
import json
from openpyxl import load_workbook
from xml.etree.ElementTree import Element, SubElement, ElementTree
# Convert the json in json_to_excel.json file to xlsx with Name, Age, City, Hobbies as header and its data
# Read from the generated excel and convert and write it to XML

def json_to_excel(input_file_path: str, excel_file_path: str):
    with open(input_file_path, 'r') as file:
        data = json.load(file)
        
    wb = Workbook()
    ws = wb.active
    ws.title = 'data'
    
    # TODO write the header to the spreadsheet
    for name, details in data.items():
        # TODO append all the data to the spreadsheet
        continue
    
    wb.save(excel_file_path)


json_file_path = "json_to_excel.json"
json_excel_path = "output_json.xlsx"


json_to_excel(json_file_path, json_excel_path)
print(f"Data from {json_file_path} has been written to {json_excel_path}.")


def excel_to_xml(excel_file_path: str, xml_file_path: str):
    # Load the workbook and select the active sheet
    wb = load_workbook(excel_file_path)
    ws = wb.active
    
    # TODO Create the root element
    root = None
    
    # Iterate through rows, skipping the header
    for row in ws.iter_rows(min_row=2, values_only=True):
        person = SubElement(root, 'Person')
        
        # TODO Map values to XML structure
    
    # Write XML to file
    tree = ElementTree(root)
    tree.write(xml_file_path, encoding='utf-8', xml_declaration=True)


excel_file_path = "output_json.xlsx"
xml_file_path = "output_xml.xml"

# Convert Excel to XML
excel_to_xml(excel_file_path, xml_file_path)
print(f"Data from {excel_file_path} has been written to {xml_file_path}.")